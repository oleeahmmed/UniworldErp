from django.shortcuts import render
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from uniworlderp.models import SalesOrder, CustomerVendor, Product, SalesEmployee, SalesOrderItem, ReturnSalesItem
from django.db.models import Sum, F, Q
from datetime import datetime, date
from django.utils import timezone
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io

class SalesOrderReportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for generating and displaying sales order reports with customer-wise analysis."""
    template_name = 'reports/sales_order_report.html'
    permission_required = 'uniworlderp.view_salesorder'
    
    def get(self, request, *args, **kwargs):
        """Display sales order report form and results."""
        # Get filter parameters
        customer_id = request.GET.get('customer', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        
        # Validate date range if both dates are provided
        if start_date and end_date:
            try:
                from datetime import datetime
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date() if isinstance(start_date, str) else start_date
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date() if isinstance(end_date, str) else end_date
                
                if end_date_obj < start_date_obj:
                    error_message = "End date must be after start date."
                    customers = CustomerVendor.objects.filter(entity_type='customer', owner=request.user).order_by('name')
                    return render(request, self.template_name, {
                        'customers': customers,
                        'selected_customer': customer_id,
                        'start_date': start_date,
                        'end_date': end_date,
                        'error': error_message,
                        'sales_orders': [],
                        'total_orders': 0,
                        'total_gross_qty': 0,
                        'total_returned_qty': 0,
                        'total_net_qty': 0,
                        'total_gross_amount': 0,
                        'total_returned_amount': 0,
                        'total_net_amount': 0,
                    })
            except (ValueError, TypeError):
                error_message = "Invalid date format."
                customers = CustomerVendor.objects.filter(entity_type='customer', owner=request.user).order_by('name')
                return render(request, self.template_name, {
                    'customers': customers,
                    'selected_customer': customer_id,
                    'start_date': start_date,
                    'end_date': end_date,
                    'error': error_message,
                    'sales_orders': [],
                    'total_orders': 0,
                    'total_gross_qty': 0,
                    'total_returned_qty': 0,
                    'total_net_qty': 0,
                    'total_gross_amount': 0,
                    'total_returned_amount': 0,
                    'total_net_amount': 0,
                })
        
        # Get all customers for the filter dropdown
        customers = CustomerVendor.objects.filter(entity_type='customer', owner=request.user).order_by('name')
        
        # Start with all sales orders
        sales_orders = SalesOrder.objects.filter(owner=request.user).select_related(
            'customer', 'sales_employee'
        ).prefetch_related('order_items__product').order_by('-order_date')
        
        # Apply customer filter
        if customer_id:
            sales_orders = sales_orders.filter(customer_id=customer_id)
        
        # Apply date range filter
        if start_date:
            sales_orders = sales_orders.filter(order_date__gte=start_date)
        if end_date:
            sales_orders = sales_orders.filter(order_date__lte=end_date)
        
        # Get order IDs for return data query
        order_ids = list(sales_orders.values_list('id', flat=True))
        
        # Query return data grouped by sales order
        returns_qs = ReturnSalesItem.objects.filter(
            sales_order_item__sales_order__id__in=order_ids
        )
        
        # Apply date range filter to returns using return_date
        if start_date and end_date:
            returns_qs = returns_qs.filter(
                return_sales__return_date__range=[start_date, end_date]
            )
        elif start_date:
            returns_qs = returns_qs.filter(
                return_sales__return_date__gte=start_date
            )
        elif end_date:
            returns_qs = returns_qs.filter(
                return_sales__return_date__lte=end_date
            )
        
        # Aggregate returns by sales order
        returns_aggregated = returns_qs.values(
            'sales_order_item__sales_order__id'
        ).annotate(
            returned_qty=Sum('quantity'),
            returned_amount=Sum('total')
        )
        
        # Build returns_by_order dictionary
        returns_by_order = {}
        for item in returns_aggregated:
            order_id = item['sales_order_item__sales_order__id']
            returns_by_order[order_id] = {
                'qty': item['returned_qty'] or 0,
                'amount': item['returned_amount'] or 0
            }
        
        # Attach return data to each order and calculate net values
        total_gross_qty = 0
        total_returned_qty = 0
        total_net_qty = 0
        total_gross_amount = 0
        total_returned_amount = 0
        total_net_amount = 0
        
        for order in sales_orders:
            # Get return data for this order
            returns = returns_by_order.get(order.id, {'qty': 0, 'amount': 0})
            order.returned_qty = returns['qty']
            order.returned_amount = returns['amount']
            
            # Calculate gross quantity from order items
            order.gross_qty = sum(item.quantity for item in order.order_items.all())
            
            # Calculate net quantity
            order.net_qty = order.gross_qty - order.returned_qty
            
            # Gross amount is the order's total_amount
            order.gross_amount = order.total_amount
            
            # Calculate net amount
            order.net_amount = order.gross_amount - order.returned_amount
            
            # Accumulate totals
            total_gross_qty += order.gross_qty
            total_returned_qty += order.returned_qty
            total_net_qty += order.net_qty
            total_gross_amount += order.gross_amount
            total_returned_amount += order.returned_amount
            total_net_amount += order.net_amount
        
        # Calculate summary data
        total_orders = sales_orders.count()
        
        # Get customer information for header if customer is selected
        customer_info = None
        if customer_id:
            try:
                customer_info = CustomerVendor.objects.get(id=customer_id, owner=request.user)
            except CustomerVendor.DoesNotExist:
                pass
        
        # Prepare context
        context = {
            'sales_orders': sales_orders,
            'customers': customers,
            'selected_customer': customer_id,
            'start_date': start_date,
            'end_date': end_date,
            'total_orders': total_orders,
            'total_gross_qty': total_gross_qty,
            'total_returned_qty': total_returned_qty,
            'total_net_qty': total_net_qty,
            'total_gross_amount': total_gross_amount,
            'total_returned_amount': total_returned_amount,
            'total_net_amount': total_net_amount,
            'customer_info': customer_info,
        }
        
        return render(request, self.template_name, context)

class SalesOrderReportPrintView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for printing sales order reports."""
    template_name = 'reports/sales_order_report_print.html'
    permission_required = 'uniworlderp.view_salesorder'
    
    def get(self, request, *args, **kwargs):
        """Display printable sales order report."""
        # Get filter parameters
        customer_id = request.GET.get('customer', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        
        # Get customer information for header if customer is selected
        customer_info = None
        if customer_id:
            try:
                customer_info = CustomerVendor.objects.get(id=customer_id, owner=request.user)
            except CustomerVendor.DoesNotExist:
                pass
        
        # Start with all sales orders
        sales_orders = SalesOrder.objects.filter(owner=request.user).select_related(
            'customer', 'sales_employee'
        ).prefetch_related('order_items__product').order_by('-order_date')
        
        # Apply filters
        if customer_id:
            sales_orders = sales_orders.filter(customer_id=customer_id)
        if start_date:
            sales_orders = sales_orders.filter(order_date__gte=start_date)
        if end_date:
            sales_orders = sales_orders.filter(order_date__lte=end_date)
        
        # Get order IDs for return data query
        order_ids = list(sales_orders.values_list('id', flat=True))
        
        # Query return data grouped by sales order
        returns_qs = ReturnSalesItem.objects.filter(
            sales_order_item__sales_order__id__in=order_ids
        )
        
        # Apply date range filter to returns using return_date
        if start_date and end_date:
            returns_qs = returns_qs.filter(
                return_sales__return_date__range=[start_date, end_date]
            )
        elif start_date:
            returns_qs = returns_qs.filter(
                return_sales__return_date__gte=start_date
            )
        elif end_date:
            returns_qs = returns_qs.filter(
                return_sales__return_date__lte=end_date
            )
        
        # Aggregate returns by sales order
        returns_aggregated = returns_qs.values(
            'sales_order_item__sales_order__id'
        ).annotate(
            returned_qty=Sum('quantity'),
            returned_amount=Sum('total')
        )
        
        # Build returns_by_order dictionary
        returns_by_order = {}
        for item in returns_aggregated:
            order_id = item['sales_order_item__sales_order__id']
            returns_by_order[order_id] = {
                'qty': item['returned_qty'] or 0,
                'amount': item['returned_amount'] or 0
            }
        
        # Attach return data to each order and calculate net values
        total_gross_qty = 0
        total_returned_qty = 0
        total_net_qty = 0
        total_gross_amount = 0
        total_returned_amount = 0
        total_net_amount = 0
        
        for order in sales_orders:
            # Get return data for this order
            returns = returns_by_order.get(order.id, {'qty': 0, 'amount': 0})
            order.returned_qty = returns['qty']
            order.returned_amount = returns['amount']
            
            # Calculate gross quantity from order items
            order.gross_qty = sum(item.quantity for item in order.order_items.all())
            
            # Calculate net quantity
            order.net_qty = order.gross_qty - order.returned_qty
            
            # Gross amount is the order's total_amount
            order.gross_amount = order.total_amount
            
            # Calculate net amount
            order.net_amount = order.gross_amount - order.returned_amount
            
            # Accumulate totals
            total_gross_qty += order.gross_qty
            total_returned_qty += order.returned_qty
            total_net_qty += order.net_qty
            total_gross_amount += order.gross_amount
            total_returned_amount += order.returned_amount
            total_net_amount += order.net_amount
        
        # Calculate summary data
        total_orders = sales_orders.count()
        
        context = {
            'sales_orders': sales_orders,
            'customer_info': customer_info,
            'start_date': start_date,
            'end_date': end_date,
            'total_orders': total_orders,
            'total_gross_qty': total_gross_qty,
            'total_returned_qty': total_returned_qty,
            'total_net_qty': total_net_qty,
            'total_gross_amount': total_gross_amount,
            'total_returned_amount': total_returned_amount,
            'total_net_amount': total_net_amount,
        }
        
        return render(request, self.template_name, context)

class SalesOrderReportExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for exporting sales order reports to Excel."""
    permission_required = 'uniworlderp.view_salesorder'
    
    def get(self, request, *args, **kwargs):
        """Export sales order report to Excel."""
        # Get filter parameters
        customer_id = request.GET.get('customer', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        
        # Get customer information for header if customer is selected
        customer_info = None
        if customer_id:
            try:
                customer_info = CustomerVendor.objects.get(id=customer_id, owner=request.user)
            except CustomerVendor.DoesNotExist:
                pass
        
        # Start with all sales orders
        sales_orders = SalesOrder.objects.filter(owner=request.user).select_related(
            'customer', 'sales_employee'
        ).prefetch_related('order_items__product').order_by('-order_date')
        
        # Apply filters
        if customer_id:
            sales_orders = sales_orders.filter(customer_id=customer_id)
        if start_date:
            sales_orders = sales_orders.filter(order_date__gte=start_date)
        if end_date:
            sales_orders = sales_orders.filter(order_date__lte=end_date)
        
        # Get order IDs for return data query
        order_ids = list(sales_orders.values_list('id', flat=True))
        
        # Query return data grouped by sales order
        returns_qs = ReturnSalesItem.objects.filter(
            sales_order_item__sales_order__id__in=order_ids
        )
        
        # Apply date range filter to returns using return_date
        if start_date and end_date:
            returns_qs = returns_qs.filter(
                return_sales__return_date__range=[start_date, end_date]
            )
        elif start_date:
            returns_qs = returns_qs.filter(
                return_sales__return_date__gte=start_date
            )
        elif end_date:
            returns_qs = returns_qs.filter(
                return_sales__return_date__lte=end_date
            )
        
        # Aggregate returns by sales order
        returns_aggregated = returns_qs.values(
            'sales_order_item__sales_order__id'
        ).annotate(
            returned_qty=Sum('quantity'),
            returned_amount=Sum('total')
        )
        
        # Build returns_by_order dictionary
        returns_by_order = {}
        for item in returns_aggregated:
            order_id = item['sales_order_item__sales_order__id']
            returns_by_order[order_id] = {
                'qty': item['returned_qty'] or 0,
                'amount': item['returned_amount'] or 0
            }
        
        # Attach return data to each order and calculate net values
        for order in sales_orders:
            # Get return data for this order
            returns = returns_by_order.get(order.id, {'qty': 0, 'amount': 0})
            order.returned_qty = returns['qty']
            order.returned_amount = returns['amount']
            
            # Calculate gross quantity from order items
            order.gross_qty = sum(item.quantity for item in order.order_items.all())
            
            # Calculate net quantity
            order.net_qty = order.gross_qty - order.returned_qty
            
            # Gross amount is the order's total_amount
            order.gross_amount = order.total_amount
            
            # Calculate net amount
            order.net_amount = order.gross_amount - order.returned_amount
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Order Report"
        
        # Add customer information header if customer is selected
        if customer_info:
            ws.cell(row=1, column=1, value=f"Customer: {customer_info.name}")
            ws.cell(row=2, column=1, value=f"Mobile: {customer_info.phone_number}")
            ws.cell(row=3, column=1, value=f"Address: {customer_info.address or 'N/A'}")
            ws.cell(row=4, column=1, value="")
            header_row = 5
        else:
            header_row = 1
        
        # Add headers
        headers = [
            'SL', 'Order Date', 'Order ID', 'Customer', 
            'Gross Sold', 'Qty Returned', 'Net Qty', 
            'Gross Amount', 'Return Amount', 'Net Amount', 
            'Sales Employee'
        ]
        
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        row_num = header_row + 1
        for i, order in enumerate(sales_orders, 1):
            row = [
                i,
                order.order_date.strftime('%Y-%m-%d') if order.order_date else '',
                order.id,
                order.customer.name if order.customer else '',
                order.gross_qty,
                order.returned_qty or 0,
                order.net_qty,
                float(order.gross_amount),
                float(order.returned_amount or 0),
                float(order.net_amount),
                order.sales_employee.full_name if order.sales_employee else '',
            ]
            for col, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col, value=value)
            row_num += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=sales_order_report.xlsx'
        
        return response
