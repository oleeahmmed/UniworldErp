from django.shortcuts import render
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from uniworlderp.models import SalesOrder, CustomerVendor, Product, SalesEmployee, SalesOrderItem, StockTransaction, ReturnSalesItem
from django.db.models import Sum, F, ExpressionWrapper, DecimalField, Q, Value, IntegerField
from django.db.models.functions import Coalesce
from datetime import datetime, timedelta, time
from django.utils import timezone
from django.utils.dateparse import parse_date
import pytz
from uniworlderp.forms import StockReportForm
from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
from decimal import Decimal


# Minimum allowed date for any stock report queries
MIN_STOCK_DATE = timezone.make_aware(datetime(2025, 7, 27))

class ReportView(LoginRequiredMixin, View):
    template_name = 'reports/report.html'

    def get(self, request):
        # Fetch data for filters
        customers = CustomerVendor.objects.filter(entity_type='customer').order_by('name')
        products = Product.objects.all().order_by('name')
        sales_employees = SalesEmployee.objects.all().order_by('full_name')

        # Initialize empty summaries for tabs
        customer_summary = []
        product_summary = []
        sales_employee_summary = []
        date_summary = []

        # Render the template with context
        return render(request, self.template_name, {
            'customers': customers,
            'products': products,
            'sales_employees': sales_employees,
            'customer_summary': customer_summary,
            'product_summary': product_summary,
            'sales_employee_summary': sales_employee_summary,
            'date_summary': date_summary,
        })

    def post(self, request):
        # Handle form submission and filter data
        customer_id = request.POST.get('customer')
        product_id = request.POST.get('product')
        sales_employee_id = request.POST.get('sales_employee')
        
        # Set default start and end dates
        today = timezone.now().date()
        first_day_of_month = today.replace(day=1)
        
        start_date = request.POST.get('start_date', first_day_of_month)
        end_date = request.POST.get('end_date', today)
        
        # Validate date range if both dates are provided
        if start_date and end_date:
            try:
                from datetime import datetime
                start_date_obj = datetime.strptime(str(start_date), '%Y-%m-%d').date() if isinstance(start_date, str) else start_date
                end_date_obj = datetime.strptime(str(end_date), '%Y-%m-%d').date() if isinstance(end_date, str) else end_date
                
                if end_date_obj < start_date_obj:
                    error_message = "End date must be after start date."
                    return render(request, self.template_name, {
                        'customers': CustomerVendor.objects.filter(entity_type='customer').order_by('name'),
                        'products': Product.objects.all().order_by('name'),
                        'sales_employees': SalesEmployee.objects.all().order_by('full_name'),
                        'customer_summary': [],
                        'product_summary': [],
                        'sales_employee_summary': [],
                        'date_summary': [],
                        'error': error_message
                    })
            except (ValueError, TypeError):
                error_message = "Invalid date format."
                return render(request, self.template_name, {
                    'customers': CustomerVendor.objects.filter(entity_type='customer').order_by('name'),
                    'products': Product.objects.all().order_by('name'),
                    'sales_employees': SalesEmployee.objects.all().order_by('full_name'),
                    'customer_summary': [],
                    'product_summary': [],
                    'sales_employee_summary': [],
                    'date_summary': [],
                    'error': error_message
                })

        # Query SalesOrderItem directly for unified item-level report format
        # This provides consistent display across all filter combinations
        items = SalesOrderItem.objects.select_related(
            'sales_order__customer',
            'sales_order__sales_employee',
            'product'
        )
        
        # Apply filters on item-level
        if customer_id:
            items = items.filter(sales_order__customer_id=customer_id)
        if product_id:
            items = items.filter(product_id=product_id)
        if sales_employee_id:
            items = items.filter(sales_order__sales_employee_id=sales_employee_id)
        if start_date and end_date:
            items = items.filter(sales_order__order_date__range=[start_date, end_date])
        
        # Order results by date (descending) and sales order id
        items = items.order_by('-sales_order__order_date', 'sales_order__id')

        # Query ReturnSalesItem data grouped by sales_order_item
        returns_qs = ReturnSalesItem.objects.select_related(
            'sales_order_item__sales_order__customer',
            'sales_order_item__sales_order__sales_employee',
            'sales_order_item__product',
            'return_sales'
        )
        
        # Apply same filters to returns
        if customer_id:
            returns_qs = returns_qs.filter(
                sales_order_item__sales_order__customer_id=customer_id
            )
        if product_id:
            returns_qs = returns_qs.filter(
                sales_order_item__product_id=product_id
            )
        if sales_employee_id:
            returns_qs = returns_qs.filter(
                sales_order_item__sales_order__sales_employee_id=sales_employee_id
            )
        if start_date and end_date:
            returns_qs = returns_qs.filter(
                return_sales__return_date__range=[start_date, end_date]
            )
        
        # Group returns by sales_order_item_id
        returns_by_item = returns_qs.values('sales_order_item_id').annotate(
            returned_qty=Sum('quantity'),
            returned_amount=Sum('total')
        )
        
        # Build dictionary for quick lookup
        returns_dict = {
            r['sales_order_item_id']: {
                'qty': r['returned_qty'] or 0,
                'amount': r['returned_amount'] or 0
            }
            for r in returns_by_item
        }
        
        # Attach return data and calculate net values for each item
        items_with_data = []
        for item in items:
            # Attach return data for this specific item
            returns = returns_dict.get(item.id, {'qty': 0, 'amount': 0})
            item.returned_qty = returns['qty']
            item.returned_amount = returns['amount']
            
            # Calculate gross amount (before item-level discount)
            item.gross_amount = item.quantity * item.unit_price
            
            # Calculate net values for this item
            # Use item.total directly (already has item-level discount applied)
            item.net_qty = item.quantity - item.returned_qty
            item.net_amount = item.total - item.returned_amount
            
            items_with_data.append(item)
        
        # Aggregate total returned quantity and amount using Sum()
        returns_aggregated = returns_qs.aggregate(
            total_returned_qty=Sum('quantity'),
            total_returned_amount=Sum('total')
        )
        
        # Handle None values from aggregate (default to 0)
        returned_qty = returns_aggregated['total_returned_qty'] or 0
        returned_amount = returns_aggregated['total_returned_amount'] or 0
        
        # Calculate gross_qty from items queryset
        gross_qty = sum(item.quantity for item in items_with_data)
        
        # Calculate net_qty = gross_qty - returned_qty
        net_qty = gross_qty - returned_qty
        
        # Calculate gross_amount from items (using item.total which has item-level discount)
        gross_amount = sum(item.total for item in items_with_data)
        
        # Calculate net_amount = gross_amount - returned_amount
        net_amount = gross_amount - returned_amount
        
        # Calculate summary amounts with breakdown (gross, discount, return, net)
        # Group by customer
        customer_totals = {}
        for item in items_with_data:
            customer_name = item.sales_order.customer.name
            if customer_name not in customer_totals:
                customer_totals[customer_name] = {
                    'gross_amount': Decimal('0.00'),
                    'discount_amount': Decimal('0.00'),
                    'return_amount': Decimal('0.00'),
                    'net_amount': Decimal('0.00')
                }
            customer_totals[customer_name]['gross_amount'] += item.gross_amount
            customer_totals[customer_name]['discount_amount'] += (item.total_discount or Decimal('0.00'))
            customer_totals[customer_name]['return_amount'] += item.returned_amount
            customer_totals[customer_name]['net_amount'] += item.net_amount
        
        customer_summary = [
            {
                'sales_order__customer__name': k,
                'gross_amount': v['gross_amount'],
                'discount_amount': v['discount_amount'],
                'return_amount': v['return_amount'],
                'net_amount': v['net_amount']
            }
            for k, v in sorted(customer_totals.items())
        ]
        
        # Group by product
        product_totals = {}
        for item in items_with_data:
            product_name = item.product.name
            if product_name not in product_totals:
                product_totals[product_name] = {
                    'gross_amount': Decimal('0.00'),
                    'discount_amount': Decimal('0.00'),
                    'return_amount': Decimal('0.00'),
                    'net_amount': Decimal('0.00')
                }
            product_totals[product_name]['gross_amount'] += item.gross_amount
            product_totals[product_name]['discount_amount'] += (item.total_discount or Decimal('0.00'))
            product_totals[product_name]['return_amount'] += item.returned_amount
            product_totals[product_name]['net_amount'] += item.net_amount
        
        product_summary = [
            {
                'product__name': k,
                'gross_amount': v['gross_amount'],
                'discount_amount': v['discount_amount'],
                'return_amount': v['return_amount'],
                'net_amount': v['net_amount']
            }
            for k, v in sorted(product_totals.items())
        ]
        
        # Group by sales employee
        employee_totals = {}
        for item in items_with_data:
            employee_name = item.sales_order.sales_employee.full_name if item.sales_order.sales_employee else 'Unassigned'
            if employee_name not in employee_totals:
                employee_totals[employee_name] = {
                    'gross_amount': Decimal('0.00'),
                    'discount_amount': Decimal('0.00'),
                    'return_amount': Decimal('0.00'),
                    'net_amount': Decimal('0.00')
                }
            employee_totals[employee_name]['gross_amount'] += item.gross_amount
            employee_totals[employee_name]['discount_amount'] += (item.total_discount or Decimal('0.00'))
            employee_totals[employee_name]['return_amount'] += item.returned_amount
            employee_totals[employee_name]['net_amount'] += item.net_amount
        
        sales_employee_summary = [
            {
                'sales_order__sales_employee__full_name': k,
                'gross_amount': v['gross_amount'],
                'discount_amount': v['discount_amount'],
                'return_amount': v['return_amount'],
                'net_amount': v['net_amount']
            }
            for k, v in sorted(employee_totals.items())
        ]

        # Group by date
        date_totals = {}
        for item in items_with_data:
            order_date = item.sales_order.order_date
            if order_date not in date_totals:
                date_totals[order_date] = {
                    'gross_amount': Decimal('0.00'),
                    'discount_amount': Decimal('0.00'),
                    'return_amount': Decimal('0.00'),
                    'net_amount': Decimal('0.00')
                }
            date_totals[order_date]['gross_amount'] += item.gross_amount
            date_totals[order_date]['discount_amount'] += (item.total_discount or Decimal('0.00'))
            date_totals[order_date]['return_amount'] += item.returned_amount
            date_totals[order_date]['net_amount'] += item.net_amount
        
        date_summary = [
            {
                'sales_order__order_date': k,
                'gross_amount': v['gross_amount'],
                'discount_amount': v['discount_amount'],
                'return_amount': v['return_amount'],
                'net_amount': v['net_amount']
            }
            for k, v in sorted(date_totals.items())
        ]

        # Render the template with filtered item-level data and summaries
        return render(request, self.template_name, {
            'report_items': items_with_data,
            'customers': CustomerVendor.objects.filter(entity_type='customer').order_by('name'),
            'products': Product.objects.all().order_by('name'),
            'sales_employees': SalesEmployee.objects.all().order_by('full_name'),
            'customer_summary': customer_summary,
            'product_summary': product_summary,
            'sales_employee_summary': sales_employee_summary,
            'date_summary': date_summary,
            'start_date': start_date,
            'end_date': end_date,
            'gross_qty': gross_qty,
            'returned_qty': returned_qty,
            'net_qty': net_qty,
            'gross_amount': gross_amount,
            'returned_amount': returned_amount,
            'net_amount': net_amount,
        })

    def get_product_transactions(self, product, start_date=None, end_date=None):
        """Get stock transactions (IN/OUT/RET/ADJ) for a product within an optional date range."""
        qs = StockTransaction.objects.filter(product=product)

        # Support dates as strings (YYYY-MM-DD) or date objects
        if start_date and end_date:
            qs = qs.filter(transaction_date__date__range=[start_date, end_date])

        qs = qs.order_by('transaction_date', 'id')

        txns = []
        for t in qs:
            ttype = t.transaction_type
            # Signed quantity for display/running math (+ for IN/RET, - for OUT). ADJ sets absolute stock
            if ttype in ('IN', 'RET'):
                signed_qty = t.quantity
            elif ttype == 'OUT':
                signed_qty = -t.quantity
            else:  # 'ADJ'
                signed_qty = 0

            txns.append({
                'datetime': getattr(t, 'transaction_date', None),
                'type': ttype,
                'type_label': t.get_transaction_type_display() if hasattr(t, 'get_transaction_type_display') else ttype,
                'quantity': t.quantity if ttype in ('IN', 'RET') else t.quantity,  # Raw quantity for display
                'signed_qty': signed_qty,  # Signed quantity for calculations
                'in_qty': signed_qty if signed_qty > 0 else 0,
                'out_qty': abs(signed_qty) if signed_qty < 0 else 0,
                'reference': getattr(t, 'reference', ''),
                'previous_stock': getattr(t, 'previous_stock', None),
                'current_stock': getattr(t, 'current_stock', None),
            })

        return txns

    def calculate_summary(self, transactions):
        """Calculate opening, total received, total issued, returns, and closing stock from transaction rows."""
        if not transactions:
            return {
                'opening_stock': 0,
                'total_in': 0,
                'total_issued': 0,
                'total_returned': 0,
                'total_received': 0,
                'closing_stock': 0,
            }

        opening_stock = transactions[0].get('previous_stock') or 0
        closing_stock = transactions[-1].get('current_stock') or opening_stock

        # Separate calculations for each transaction type
        total_in = sum(t['quantity'] for t in transactions if t.get('type') == 'IN')
        total_out = sum(abs(t['quantity']) for t in transactions if t.get('type') == 'OUT')
        total_returned = sum(t['quantity'] for t in transactions if t.get('type') == 'RET')
        
        # Legacy calculation for backward compatibility
        total_received = sum(t['quantity'] for t in transactions if t['quantity'] > 0)
        total_issued = sum(abs(t['quantity']) for t in transactions if t['quantity'] < 0)

        return {
            'opening_stock': opening_stock,
            'total_in': total_in,
            'total_issued': total_out,
            'total_returned': total_returned,
            'total_received': total_received,
            'closing_stock': closing_stock,
        }


class StockReportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for generating and displaying stock reports."""
    template_name = 'reports/stock_report.html'
    permission_required = 'uniworlderp.view_product'

    def get(self, request, *args, **kwargs):
        form = StockReportForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        form = StockReportForm(request.POST)
        if form.is_valid():
            report_data, context = self.generate_report_data(form)
            
            # --- Start Debug Logging ---
            import logging
            logger = logging.getLogger(__name__)
            logger.setLevel(logging.DEBUG)
            
            # Create a handler that writes to stderr
            handler = logging.StreamHandler()
            handler.setLevel(logging.DEBUG)
            formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            
            # Add the handler to the logger
            if not logger.handlers:
                logger.addHandler(handler)
            
            logger.debug("--- Stock Report Data ---")
            for item in report_data:
                logger.debug(item)
            logger.debug("-------------------------")
            # --- End Debug Logging ---
            
            context['form'] = form
            context['report_data'] = report_data

            # If a single product is selected, include detailed transactions in the same page
            product_obj = form.cleaned_data.get('product_id')
            if product_obj:
                # Determine date range as dates for filtering
                date_range = form.cleaned_data.get('date_range')
                start_date = form.cleaned_data.get('start_date')
                end_date = form.cleaned_data.get('end_date')
                if date_range == 'today':
                    # Use today's date for both bounds
                    today_bd = timezone.localtime().date()
                    start_date = today_bd
                    end_date = today_bd

                report_view = ReportView()
                transactions = report_view.get_product_transactions(product_obj, start_date, end_date)
                summary = report_view.calculate_summary(transactions)

                context.update({
                    'single_product': product_obj,
                    'single_product_transactions': transactions,
                    'single_product_summary': summary,
                    'single_product_start_date': start_date,
                    'single_product_end_date': end_date,
                })
            return render(request, self.template_name, context)
        return render(request, self.template_name, {'form': form})

    @staticmethod
    def generate_report_data(form):
        product_id = form.cleaned_data.get('product_id')
        date_range = form.cleaned_data.get('date_range')
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')

        now = timezone.now()
        bdt = pytz.timezone('Asia/Dhaka')
        now_bdt = now.astimezone(bdt)
        
        # --- Date Handling ---
        if date_range == 'today':
            start_dt = bdt.localize(datetime.combine(now_bdt.date(), time.min))
            end_dt = now_bdt
            report_date_display = f"{now_bdt.date().strftime('%d/%m/%Y')}"
        else: # custom range
            start_dt = bdt.localize(datetime.combine(start_date, time.min))
            if end_date == now_bdt.date():
                end_dt = now_bdt
            else:
                end_dt = bdt.localize(datetime.combine(end_date, time.max))
            
            if start_date == end_date:
                report_date_display = f"{start_date.strftime('%d/%m/%Y')}"
            else:
                report_date_display = f"{start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}"

        products = Product.objects.filter(is_active=True)
        if product_id:
            products = products.filter(pk=product_id.pk)

        report_results = []
        with transaction.atomic():
            for i, product in enumerate(products, 1):
                # --- Stock Calculation Logic ---
                
                # 1. Get movements within the date range
                transactions_in_range = StockTransaction.objects.filter(
                    product=product,
                    transaction_date__gte=start_dt,
                    transaction_date__lte=end_dt
                ).aggregate(
                    received=Coalesce(Sum('quantity', filter=Q(transaction_type__in=['IN', 'RET'])), 0, output_field=IntegerField()),
                    issued=Coalesce(Sum('quantity', filter=Q(transaction_type='OUT')), 0, output_field=IntegerField()),
                )
                received_qty = transactions_in_range['received']
                issued_qty = transactions_in_range['issued']
                
                # 2. Determine Closing Stock
                # Start with the product's current stock as the most reliable value
                closing_stock = product.stock_quantity or 0

                # 3. Determine Opening Stock
                # Opening Stock = Closing Stock - Received Qty + Issued Qty
                opening_stock = closing_stock - received_qty + issued_qty

                remarks = "Order Required" if closing_stock <= product.reorder_level else ""

                report_results.append({
                    'sl': i,
                    'product_name': product.name,
                    'product_code': product.sku,
                    'unit': product.get_unit_display(),
                    'opening_stock': opening_stock,
                    'received_qty': received_qty,
                    'issued_qty': issued_qty,
                    'closing_stock': closing_stock,
                    'remarks': remarks,
                })

        # --- Context for Template ---
        report_start_time = start_dt.strftime('%d/%m/%Y %I:%M %p')
        report_end_time = end_dt.strftime('%d/%m/%Y %I:%M %p')

        context = {
            'report_generated_at_formatted': now_bdt.strftime('%d/%m/%Y %I:%M %p'),
            'report_date_display': report_date_display,
            'report_start_time': report_start_time,
            'report_end_time': report_end_time,
            'get_params': form.data.urlencode() if hasattr(form.data, 'urlencode') else ''
        }
        return report_results, context


class SingleProductReportPrintView(LoginRequiredMixin, View):
    """View for printing single product transaction reports."""
    template_name = 'reports/single_product_transaction_report.html'

    def get(self, request, *args, **kwargs):
        product_id = request.GET.get('product_id')
        start_date_raw = request.GET.get('start_date')
        end_date_raw = request.GET.get('end_date')
        # Normalize dates: accept YYYY-MM-DD or other human strings
        start_date = parse_date(start_date_raw) if start_date_raw else None
        end_date = parse_date(end_date_raw) if end_date_raw else None
        if start_date is None and start_date_raw:
            # fallback: try common format like 'Aug. 11, 2025' without the dot
            try:
                from datetime import datetime
                start_date = datetime.strptime(start_date_raw.replace('.', ''), '%b %d, %Y').date()
            except Exception:
                start_date = None
        if end_date is None and end_date_raw:
            try:
                from datetime import datetime
                end_date = datetime.strptime(end_date_raw.replace('.', ''), '%b %d, %Y').date()
            except Exception:
                end_date = None
        
        if not product_id:
            return render(request, self.template_name, {'error': 'Product ID is required.'})
        
        try:
            product = Product.objects.get(id=product_id, is_active=True)
            report_view = ReportView()
            transactions = report_view.get_product_transactions(product, start_date, end_date)
            summary = report_view.calculate_summary(transactions)
            
            context = {
                'product': product,
                'transactions': transactions,
                'summary': summary,
                'start_date': start_date,
                'end_date': end_date,
                'print_view': True,
            }
            return render(request, self.template_name, context)
        except Product.DoesNotExist:
            return render(request, self.template_name, {'error': 'Product not found or inactive.'})


class SingleProductStockReportPrintView(LoginRequiredMixin, View):
    """View for printing single product stock reports with proper format."""
    template_name = 'reports/single_product_stock_report_print.html'

    def get(self, request, *args, **kwargs):
        product_id = request.GET.get('product_id')
        start_date_raw = request.GET.get('start_date')
        end_date_raw = request.GET.get('end_date')
        
        # Parse dates
        start_date = parse_date(start_date_raw) if start_date_raw else None
        end_date = parse_date(end_date_raw) if end_date_raw else None
        
        if not product_id:
            return render(request, self.template_name, {'error': 'Product ID is required.'})
        
        try:
            product = Product.objects.get(id=product_id, is_active=True)
            report_view = ReportView()
            transactions = report_view.get_product_transactions(product, start_date, end_date)
            summary = report_view.calculate_summary(transactions)
            
            # Format dates for display
            now = timezone.now()
            bdt = pytz.timezone('Asia/Dhaka')
            now_bdt = now.astimezone(bdt)
            
            if start_date and end_date:
                if start_date == end_date:
                    report_date_display = f"{start_date.strftime('%d/%m/%Y')}"
                else:
                    report_date_display = f"{start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}"
                report_start_time = start_date.strftime('%d/%m/%Y %I:%M %p')
                report_end_time = end_date.strftime('%d/%m/%Y %I:%M %p')
            else:
                report_date_display = "All Transactions"
                report_start_time = "All Time"
                report_end_time = "All Time"
            
            context = {
                'product': product,
                'transactions': transactions,
                'summary': summary,
                'start_date': start_date,
                'end_date': end_date,
                'print_view': True,
                'user': request.user,
                'report_date_display': report_date_display,
                'report_start_time': report_start_time,
                'report_end_time': report_end_time,
                'report_generated_at_formatted': now_bdt.strftime('%d/%m/%Y %I:%M %p'),
            }
            return render(request, self.template_name, context)
        except Product.DoesNotExist:
            return render(request, self.template_name, {'error': 'Product not found or inactive.'})


class StockReportPrintView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for printing stock reports."""
    template_name = 'reports/stock_report_print.html'
    permission_required = 'uniworlderp.view_product'

    def get(self, request, *args, **kwargs):
        form = StockReportForm(request.GET)
        if form.is_valid():
            report_data, context = StockReportView.generate_report_data(form)
            context['report_data'] = report_data
            return render(request, self.template_name, context)
        
        return render(request, 'reports/stock_report_print.html', {'error': 'Invalid parameters for print view.'})


class CustomerReportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for generating and displaying customer reports."""
    template_name = 'reports/customer_report.html'
    permission_required = 'uniworlderp.view_customervendor'

    def get(self, request, *args, **kwargs):
        """Display full customer report (all customers, no filters)."""
        # Fetch all customers for this owner
        customers = (
            CustomerVendor.objects
            .filter(entity_type='customer')
            .order_by('name')
        )

        # Prepare report metadata similar to stock report
        now_bd = timezone.localtime()
        context = {
            'customers': customers,
            'total_customers': customers.count(),
            'user': request.user,
            'report_date_display': now_bd.strftime('%Y-%m-%d'),
            'report_generated_at_formatted': now_bd.strftime('%Y-%m-%d %H:%M:%S'),
        }

        return render(request, self.template_name, context)


class CustomerReportPrintView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for printing customer reports."""
    template_name = 'reports/customer_report_print.html'
    permission_required = 'uniworlderp.view_customervendor'

    def get(self, request, *args, **kwargs):
        """Display printable full customer report."""
        customers = (
            CustomerVendor.objects
            .filter(entity_type='customer')
            .order_by('name')
        )

        now_bd = timezone.localtime()
        context = {
            'customers': customers,
            'total_customers': customers.count(),
            'user': request.user,
            'report_date_display': now_bd.strftime('%Y-%m-%d'),
            'report_generated_at_formatted': now_bd.strftime('%Y-%m-%d %H:%M:%S'),
        }

        return render(request, self.template_name, context)


class CustomerReportExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for exporting customer reports to Excel."""
    permission_required = 'uniworlderp.view_customervendor'

    def get(self, request, *args, **kwargs):
        """Export full customer report to Excel (all customers)."""
        customers = (
            CustomerVendor.objects
            .filter(entity_type='customer')
            .order_by('name')
        )

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Customer Report"
        
        # Add headers
        headers = [
            'SL', 'Company Name', 'Phone', 'Email', 'WhatsApp',
            'Business Type', 'Address', 'Created At', 'Updated At'
        ]
        
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for i, customer in enumerate(customers, 1):
            row = [
                i,
                customer.name,
                customer.phone_number,
                customer.email or '',
                customer.whatsapp_number or '',
                customer.get_business_type_display(),
                customer.address or '',
                customer.created_at.strftime('%Y-%m-%d %H:%M:%S') if customer.created_at else '',
                customer.updated_at.strftime('%Y-%m-%d %H:%M:%S') if customer.updated_at else '',
            ]
            for col, value in enumerate(row, 1):
                ws.cell(row=i+1, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=customer_report.xlsx'
        
        return response


# CURRENTLY DISABLED: This view is disabled in favor of using the main ReportView with product filter
# The functionality is redundant - users can get the same results by using Generate Report
# with a specific product selected and leaving customer/employee filters empty
class ProductWiseReportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for generating product-wise sales reports."""
    template_name = 'reports/product_wise_report.html'
    permission_required = 'uniworlderp.view_product'

    def get(self, request, *args, **kwargs):
        # Fetch data for filters
        products = Product.objects.all().order_by('name')
        
        # Render the template with context
        return render(request, self.template_name, {
            'products': products,
        })

    def post(self, request, *args, **kwargs):
        # Get form data
        product_id = request.POST.get('product')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/x-www-form-urlencoded'
        
        # Validation
        if not product_id:
            error_message = "Please select a product for the report."
            context = {
                'products': Product.objects.all().order_by('name'),
                'error': error_message
            }
            if is_ajax:
                return render(request, 'reports/product_wise_report_partial.html', context)
            return render(request, self.template_name, context)
        
        if not start_date or not end_date:
            error_message = "Please select both start and end dates."
            context = {
                'products': Product.objects.all().order_by('name'),
                'error': error_message
            }
            if is_ajax:
                return render(request, 'reports/product_wise_report_partial.html', context)
            return render(request, self.template_name, context)
        
        # Validate date range
        try:
            from datetime import datetime
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date() if isinstance(start_date, str) else start_date
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date() if isinstance(end_date, str) else end_date
            
            if end_date_obj < start_date_obj:
                error_message = "End date must be after start date."
                context = {
                    'products': Product.objects.all().order_by('name'),
                    'error': error_message
                }
                if is_ajax:
                    return render(request, 'reports/product_wise_report_partial.html', context)
                return render(request, self.template_name, context)
        except (ValueError, TypeError):
            error_message = "Invalid date format."
            context = {
                'products': Product.objects.all().order_by('name'),
                'error': error_message
            }
            if is_ajax:
                return render(request, 'reports/product_wise_report_partial.html', context)
            return render(request, self.template_name, context)
        
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            error_message = "Product not found."
            context = {
                'products': Product.objects.all().order_by('name'),
                'error': error_message
            }
            if is_ajax:
                return render(request, 'reports/product_wise_report_partial.html', context)
            return render(request, self.template_name, context)
        
        # Get sales data for the product within date range
        sales_data = SalesOrderItem.objects.filter(
            product=product,
            sales_order__order_date__range=[start_date, end_date]
        ).select_related(
            'sales_order__customer',
            'sales_order__sales_employee'
        ).order_by('-sales_order__order_date')
        
        # Check if no data found
        if not sales_data.exists():
            error_message = f"No sales data found for '{product.name}' between {start_date} and {end_date}."
            context = {
                'products': Product.objects.all().order_by('name'),
                'error': error_message,
                'selected_product': product,
                'start_date': start_date,
                'end_date': end_date
            }
            if is_ajax:
                return render(request, 'reports/product_wise_report_partial.html', context)
            return render(request, self.template_name, context)
        
        # Query return data grouped by sales_order_item
        returns_by_item = ReturnSalesItem.objects.filter(
            sales_order_item__product=product,
            return_sales__return_date__range=[start_date, end_date]
        ).values('sales_order_item_id').annotate(
            returned_qty=Sum('quantity'),
            returned_amount=Sum('total')
        )
        
        # Build dictionary for quick lookup
        returns_dict = {
            r['sales_order_item_id']: {
                'qty': r['returned_qty'] or 0,
                'amount': r['returned_amount'] or 0
            }
            for r in returns_by_item
        }
        
        # Attach return data to each item and calculate net values
        gross_qty = 0
        gross_amount = 0
        returned_qty = 0
        returned_amount = 0
        
        for item in sales_data:
            # Get return data for this specific item
            returns = returns_dict.get(item.id, {'qty': 0, 'amount': 0})
            item.returned_qty = returns['qty']
            item.returned_amount = returns['amount']
            
            # Calculate gross amount (before discount)
            item.gross_amount = item.quantity * item.unit_price
            
            # Calculate net values for this item
            item.net_qty = item.quantity - item.returned_qty
            item.net_amount = item.total - item.returned_amount
            
            # Accumulate totals
            gross_qty += item.quantity
            gross_amount += item.total
            returned_qty += item.returned_qty
            returned_amount += item.returned_amount
        
        # Calculate net totals
        net_qty = gross_qty - returned_qty
        net_amount = gross_amount - returned_amount
        
        # Format dates for display
        now = timezone.now()
        bdt = pytz.timezone('Asia/Dhaka')
        now_bdt = now.astimezone(bdt)
        
        context = {
            'products': Product.objects.all().order_by('name'),
            'selected_product': product,
            'start_date': start_date,
            'end_date': end_date,
            'sales_data': sales_data,
            'gross_qty': gross_qty,
            'returned_qty': returned_qty,
            'net_qty': net_qty,
            'gross_amount': gross_amount,
            'returned_amount': returned_amount,
            'net_amount': net_amount,
            'user': request.user,
            'report_generated_at': now_bdt.strftime('%d/%m/%Y %I:%M %p'),
            'report_data': sales_data  # For potential future use
        }
        
        # Return partial template for AJAX requests
        if is_ajax:
            return render(request, 'reports/product_wise_report_partial.html', context)
        
        return render(request, self.template_name, context)


# CURRENTLY DISABLED: This view is disabled in favor of using the main ReportView with product filter
class ProductWiseReportPrintView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for printing product-wise sales reports."""
    template_name = 'reports/product_wise_report_print.html'
    permission_required = 'uniworlderp.view_product'

    def get(self, request, *args, **kwargs):
        product_id = request.GET.get('product')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        # Validation
        if not product_id or not start_date or not end_date:
            return render(request, self.template_name, {'error': 'Missing required parameters.'})
        
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return render(request, self.template_name, {'error': 'Product not found.'})
        
        # Get sales data
        sales_data = SalesOrderItem.objects.filter(
            product=product,
            sales_order__order_date__range=[start_date, end_date]
        ).select_related(
            'sales_order__customer',
            'sales_order__sales_employee'
        ).order_by('-sales_order__order_date')
        
        # Calculate totals
        total_qty = sum(item.quantity for item in sales_data)
        total_amount = sum(item.total for item in sales_data)
        
        # Format dates
        now = timezone.now()
        bdt = pytz.timezone('Asia/Dhaka')
        now_bdt = now.astimezone(bdt)
        
        context = {
            'product': product,
            'start_date': start_date,
            'end_date': end_date,
            'sales_data': sales_data,
            'total_qty': total_qty,
            'total_amount': total_amount,
            'user': request.user,
            'report_generated_at': now_bdt.strftime('%d/%m/%Y %I:%M %p'),
            'print_view': True
        }
        
        return render(request, self.template_name, context)


# CURRENTLY DISABLED: This view is disabled in favor of using the main ReportView with product filter
class ProductWiseReportExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for exporting product-wise sales reports to Excel."""
    permission_required = 'uniworlderp.view_product'

    def get(self, request, *args, **kwargs):
            product_id = request.GET.get('product')
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

            if not product_id or not start_date or not end_date:
                return HttpResponse('Missing required parameters', status=400)

            try:
                product = Product.objects.get(id=product_id)
            except Product.DoesNotExist:
                return HttpResponse('Product not found', status=404)

            # Get sales data
            sales_data = SalesOrderItem.objects.filter(
                product=product,
                sales_order__order_date__range=[start_date, end_date]
            ).select_related(
                'sales_order__customer',
                'sales_order__sales_employee'
            ).order_by('-sales_order__order_date')

            # Query return data grouped by sales_order_item
            returns_by_item = ReturnSalesItem.objects.filter(
                sales_order_item__product=product,
                return_sales__return_date__range=[start_date, end_date]
            ).values('sales_order_item_id').annotate(
                returned_qty=Sum('quantity'),
                returned_amount=Sum('total')
            )
            
            # Build dictionary for quick lookup
            returns_dict = {
                r['sales_order_item_id']: {
                    'qty': r['returned_qty'] or 0,
                    'amount': r['returned_amount'] or 0
                }
                for r in returns_by_item
            }
            
            # Attach return data to each item and calculate totals
            gross_qty = 0
            gross_amount = 0
            returned_qty = 0
            returned_amount = 0
            
            for item in sales_data:
                # Get return data for this specific item
                returns = returns_dict.get(item.id, {'qty': 0, 'amount': 0})
                item.returned_qty = returns['qty']
                item.returned_amount = returns['amount']
                
                # Calculate net values for this item
                item.net_qty = item.quantity - item.returned_qty
                item.net_amount = item.total - item.returned_amount
                
                # Accumulate totals
                gross_qty += item.quantity
                gross_amount += item.total
                returned_qty += item.returned_qty
                returned_amount += item.returned_amount

            # Calculate net values
            net_qty = gross_qty - returned_qty
            net_amount = gross_amount - returned_amount

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Product Wise Report"

            # Add title and product info
            ws.merge_cells('A1:L1')
            title_cell = ws.cell(row=1, column=1, value=f"Product-wise Sales Report")
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")

            ws.merge_cells('A2:L2')
            product_cell = ws.cell(row=2, column=1, value=f"Product: {product.name} | Code: {product.sku} | Unit: {product.get_unit_display()} | Price: ৳{product.price}")
            product_cell.font = Font(bold=True)
            product_cell.alignment = Alignment(horizontal="center")

            ws.merge_cells('A3:L3')
            date_cell = ws.cell(row=3, column=1, value=f"Date Range: {start_date} to {end_date}")
            date_cell.alignment = Alignment(horizontal="center")

            # Add headers - now with Gross Amount and Return Amount columns
            headers = ['Customer Name', 'Invoice No', 'Sales Order No', 'Date', 'Gross Sold', 'Qty Returned', 'Net Qty', 'Unit', 'Price', 'Gross Amount', 'Return Amount', 'Net Amount']

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Add data
            row_num = 6

            for item in sales_data:
                invoice_no = item.sales_order.invoice.id if hasattr(item.sales_order, 'invoice') and item.sales_order.invoice else 'N/A'

                row = [
                    item.sales_order.customer.name,
                    invoice_no,
                    item.sales_order.id,
                    item.sales_order.order_date.strftime('%d/%m/%Y'),
                    item.quantity,  # Gross Sold
                    item.returned_qty or 0,  # Qty Returned
                    item.net_qty,  # Net Qty
                    product.get_unit_display(),
                    float(item.unit_price),
                    float(item.total),  # Gross Amount
                    float(item.returned_amount or 0),  # Return Amount
                    float(item.net_amount)  # Net Amount
                ]

                for col, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col, value=value)

                row_num += 1

            # Add totals - using aggregated values
            total_row = row_num + 1
            ws.merge_cells(f'A{total_row}:D{total_row}')
            ws.cell(row=total_row, column=1, value='TOTALS:').font = Font(bold=True)
            ws.cell(row=total_row, column=5, value=gross_qty).font = Font(bold=True)  # Gross Sold
            ws.cell(row=total_row, column=6, value=returned_qty).font = Font(bold=True)  # Qty Returned
            ws.cell(row=total_row, column=7, value=net_qty).font = Font(bold=True)  # Net Qty
            # Skip columns 8-9 (Unit, Price)
            ws.cell(row=total_row, column=10, value=float(gross_amount)).font = Font(bold=True)  # Gross Amount
            ws.cell(row=total_row, column=11, value=float(returned_amount)).font = Font(bold=True)  # Return Amount
            ws.cell(row=total_row, column=12, value=float(net_amount)).font = Font(bold=True)  # Net Amount

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter if hasattr(column[0], 'column_letter') else None
                if column_letter:
                    for cell in column:
                        try:
                            if hasattr(cell, 'value') and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Create response
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=product_wise_report_{product.name}_{start_date}_to_{end_date}.xlsx'

            return response



# CURRENTLY DISABLED: This view is disabled in favor of using the main ReportView with customer filter
# The functionality is redundant - users can get the same results by using Generate Report
# with a specific customer selected and leaving product/employee filters empty
class CustomerWiseReportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for generating customer-wise sales reports."""
    template_name = 'reports/customer_wise_report.html'
    permission_required = 'uniworlderp.view_customervendor'

    def get(self, request, *args, **kwargs):
        # Fetch data for filters
        customers = CustomerVendor.objects.filter(entity_type='customer').order_by('name')
        
        # Render the template with context
        return render(request, self.template_name, {
            'customers': customers,
        })

    def post(self, request, *args, **kwargs):
        # Get form data
        customer_id = request.POST.get('customer')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/x-www-form-urlencoded'
        
        # Validation
        if not customer_id:
            error_message = "Please select a customer for the report."
            context = {
                'customers': CustomerVendor.objects.filter(entity_type='customer').order_by('name'),
                'error': error_message
            }
            if is_ajax:
                return render(request, 'reports/customer_wise_report_partial.html', context)
            return render(request, self.template_name, context)
        
        if not start_date or not end_date:
            error_message = "Please select both start and end dates."
            context = {
                'customers': CustomerVendor.objects.filter(entity_type='customer').order_by('name'),
                'error': error_message
            }
            if is_ajax:
                return render(request, 'reports/customer_wise_report_partial.html', context)
            return render(request, self.template_name, context)
        
        # Validate date range
        try:
            from datetime import datetime
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date() if isinstance(start_date, str) else start_date
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date() if isinstance(end_date, str) else end_date
            
            if end_date_obj < start_date_obj:
                error_message = "End date must be after start date."
                context = {
                    'customers': CustomerVendor.objects.filter(entity_type='customer').order_by('name'),
                    'error': error_message
                }
                if is_ajax:
                    return render(request, 'reports/customer_wise_report_partial.html', context)
                return render(request, self.template_name, context)
        except (ValueError, TypeError):
            error_message = "Invalid date format."
            context = {
                'customers': CustomerVendor.objects.filter(entity_type='customer').order_by('name'),
                'error': error_message
            }
            if is_ajax:
                return render(request, 'reports/customer_wise_report_partial.html', context)
            return render(request, self.template_name, context)
        
        try:
            customer = CustomerVendor.objects.get(id=customer_id, entity_type='customer')
        except CustomerVendor.DoesNotExist:
            error_message = "Customer not found."
            context = {
                'customers': CustomerVendor.objects.filter(entity_type='customer').order_by('name'),
                'error': error_message
            }
            if is_ajax:
                return render(request, 'reports/customer_wise_report_partial.html', context)
            return render(request, self.template_name, context)
        
        # Get sales orders for the customer within date range
        sales_orders = SalesOrder.objects.filter(
            customer=customer,
            order_date__range=[start_date, end_date]
        ).select_related('sales_employee').prefetch_related(
            'order_items__product'
        ).order_by('-order_date')
        
        # Check if no data found
        if not sales_orders.exists():
            error_message = f"No sales data found for '{customer.name}' between {start_date} and {end_date}."
            context = {
                'customers': CustomerVendor.objects.filter(entity_type='customer').order_by('name'),
                'error': error_message,
                'selected_customer': customer,
                'start_date': start_date,
                'end_date': end_date
            }
            if is_ajax:
                return render(request, 'reports/customer_wise_report_partial.html', context)
            return render(request, self.template_name, context)
        
        # Query return data for this customer and group by order
        order_ids = sales_orders.values_list('id', flat=True)
        returns_by_order = ReturnSalesItem.objects.filter(
            sales_order_item__sales_order__id__in=order_ids,
            return_sales__return_date__range=[start_date, end_date]
        ).values('sales_order_item__sales_order_id').annotate(
            returned_qty=Sum('quantity'),
            returned_amount=Sum('total')
        )
        
        # Build dictionary for quick lookup
        returns_dict = {
            r['sales_order_item__sales_order_id']: {
                'qty': r['returned_qty'] or 0,
                'amount': r['returned_amount'] or 0
            }
            for r in returns_by_order
        }
        
        # Calculate aggregates and attach data to each order
        gross_qty = 0
        gross_amount = 0
        total_discount_amount = 0
        returned_qty = 0
        returned_amount = 0
        
        for order in sales_orders:
            # Calculate order-level gross quantity and amount
            order.gross_qty = sum(item.quantity for item in order.order_items.all())
            # Calculate gross_amount for each item (quantity * unit_price before discount)
            order.gross_amount = sum(item.quantity * item.unit_price for item in order.order_items.all())
            order.discount_amount = sum(item.total_discount or Decimal('0.00') for item in order.order_items.all())
            
            # Get return data for this order
            returns = returns_dict.get(order.id, {'qty': 0, 'amount': 0})
            order.returned_qty = returns['qty']
            order.returned_amount = returns['amount']
            
            # Calculate net values for this order
            order.net_qty = order.gross_qty - order.returned_qty
            order.net_amount = order.total_amount - order.returned_amount
            
            # Accumulate totals
            gross_qty += order.gross_qty
            gross_amount += order.gross_amount
            total_discount_amount += order.discount_amount
            returned_qty += order.returned_qty
            returned_amount += order.returned_amount
        
        # Calculate net values
        net_qty = gross_qty - returned_qty
        net_amount = gross_amount - total_discount_amount - returned_amount
        
        # Format dates for display
        now = timezone.now()
        bdt = pytz.timezone('Asia/Dhaka')
        now_bdt = now.astimezone(bdt)
        
        context = {
            'customers': CustomerVendor.objects.filter(entity_type='customer').order_by('name'),
            'selected_customer': customer,
            'start_date': start_date,
            'end_date': end_date,
            'sales_orders': sales_orders,
            'gross_qty': gross_qty,
            'gross_amount': gross_amount,
            'total_discount_amount': total_discount_amount,
            'returned_qty': returned_qty,
            'returned_amount': returned_amount,
            'net_qty': net_qty,
            'net_amount': net_amount,
            'user': request.user,
            'report_generated_at': now_bdt.strftime('%d/%m/%Y %I:%M %p'),
        }
        
        # Return partial template for AJAX requests
        if is_ajax:
            return render(request, 'reports/customer_wise_report_partial.html', context)
        
        return render(request, self.template_name, context)


# CURRENTLY DISABLED: This view is disabled in favor of using the main ReportView with customer filter
class CustomerWiseReportPrintView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for printing customer-wise sales reports."""
    template_name = 'reports/customer_wise_report_print.html'
    permission_required = 'uniworlderp.view_customervendor'

    def get(self, request, *args, **kwargs):
        customer_id = request.GET.get('customer')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        # Validation
        if not customer_id or not start_date or not end_date:
            return render(request, self.template_name, {'error': 'Missing required parameters.'})
        
        try:
            customer = CustomerVendor.objects.get(id=customer_id, entity_type='customer')
        except CustomerVendor.DoesNotExist:
            return render(request, self.template_name, {'error': 'Customer not found.'})
        
        # Get sales orders
        sales_orders = SalesOrder.objects.filter(
            customer=customer,
            order_date__range=[start_date, end_date]
        ).select_related('sales_employee').prefetch_related(
            'order_items__product'
        ).order_by('-order_date')
        
        # Calculate totals
        total_purchase = sum(order.total_amount - order.discount for order in sales_orders)
        total_discount = sum(order.discount for order in sales_orders)
        net_amount = sum(order.total_amount for order in sales_orders)
        
        # Format dates
        now = timezone.now()
        bdt = pytz.timezone('Asia/Dhaka')
        now_bdt = now.astimezone(bdt)
        
        context = {
            'customer': customer,
            'start_date': start_date,
            'end_date': end_date,
            'sales_orders': sales_orders,
            'total_purchase': total_purchase,
            'total_discount': total_discount,
            'net_amount': net_amount,
            'user': request.user,
            'report_generated_at': now_bdt.strftime('%d/%m/%Y %I:%M %p'),
            'print_view': True
        }
        
        return render(request, self.template_name, context)


# CURRENTLY DISABLED: This view is disabled in favor of using the main ReportView with customer filter
class CustomerWiseReportExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for exporting customer-wise sales reports to Excel."""
    permission_required = 'uniworlderp.view_customervendor'

    def get(self, request, *args, **kwargs):
            customer_id = request.GET.get('customer')
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

            if not customer_id or not start_date or not end_date:
                return HttpResponse('Missing required parameters', status=400)

            try:
                customer = CustomerVendor.objects.get(id=customer_id, entity_type='customer')
            except CustomerVendor.DoesNotExist:
                return HttpResponse('Customer not found', status=404)

            # Get sales orders
            sales_orders = SalesOrder.objects.filter(
                customer=customer,
                order_date__range=[start_date, end_date]
            ).select_related('sales_employee').prefetch_related(
                'order_items__product'
            ).order_by('-order_date')

            # Query return data for this customer
            order_ids = sales_orders.values_list('id', flat=True)
            returns_data = ReturnSalesItem.objects.filter(
                sales_order_item__sales_order__id__in=order_ids,
                return_sales__return_date__range=[start_date, end_date]
            ).aggregate(
                total_returned_qty=Sum('quantity'),
                total_returned_amount=Sum('total')
            )

            # Calculate gross sales aggregates
            gross_qty = sum(
                sum(item.quantity for item in order.order_items.all())
                for order in sales_orders
            )
            gross_amount = sum(order.total_amount for order in sales_orders)

            # Extract return values with default to 0
            returned_qty = returns_data['total_returned_qty'] or 0
            returned_amount = returns_data['total_returned_amount'] or 0

            # Calculate net values
            net_qty = gross_qty - returned_qty
            net_amount = gross_amount - returned_amount

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Customer Wise Report"

            # Add title and customer info
            ws.merge_cells('A1:K1')
            title_cell = ws.cell(row=1, column=1, value=f"Customer-wise Sales Report")
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")

            ws.merge_cells('A2:K2')
            customer_cell = ws.cell(row=2, column=1, value=f"Customer Name: {customer.name}")
            customer_cell.font = Font(bold=True)
            customer_cell.alignment = Alignment(horizontal="center")

            ws.merge_cells('A3:K3')
            address_cell = ws.cell(row=3, column=1, value=f"Address: {customer.address or 'N/A'}")
            address_cell.alignment = Alignment(horizontal="center")

            ws.merge_cells('A4:K4')
            mobile_cell = ws.cell(row=4, column=1, value=f"Mobile: {customer.phone_number}")
            mobile_cell.alignment = Alignment(horizontal="center")

            ws.merge_cells('A5:K5')
            date_cell = ws.cell(row=5, column=1, value=f"Start Date: {start_date} & End Date: {end_date}")
            date_cell.alignment = Alignment(horizontal="center")

            # Add headers with six new columns
            headers = ['Order Date', 'Order ID', 'Customer', 'Product Details', 'Gross Sold', 'Qty Returned', 'Net Qty', 'Gross Amount', 'Return Amount', 'Net Amount', 'Sales Employee']

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Add data
            row_num = 8

            for order in sales_orders:
                # Create product details string
                product_details = ', '.join([f"{item.product.name} ({item.quantity})" for item in order.order_items.all()])

                # Calculate order-level gross quantity
                order_gross_qty = sum(item.quantity for item in order.order_items.all())

                # Calculate order-level gross amount
                order_gross_amount = order.total_amount

                # Query returns for this specific order
                order_returns = ReturnSalesItem.objects.filter(
                    sales_order_item__sales_order__id=order.id,
                    return_sales__return_date__range=[start_date, end_date]
                ).aggregate(
                    returned_qty=Sum('quantity'),
                    returned_amount=Sum('total')
                )

                # Extract return values with default to 0
                order_returned_qty = order_returns['returned_qty'] or 0
                order_returned_amount = order_returns['returned_amount'] or 0

                # Calculate net values for this order
                order_net_qty = order_gross_qty - order_returned_qty
                order_net_amount = order_gross_amount - order_returned_amount

                row = [
                    order.order_date.strftime('%d/%m/%Y'),
                    order.id,
                    customer.name,
                    product_details,
                    order_gross_qty,
                    order_returned_qty,
                    order_net_qty,
                    float(order_gross_amount),
                    float(order_returned_amount),
                    float(order_net_amount),
                    order.sales_employee.full_name if order.sales_employee else 'N/A'
                ]

                for col, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col, value=value)

                row_num += 1

            # Add totals
            total_row = row_num + 1
            ws.merge_cells(f'A{total_row}:D{total_row}')
            ws.cell(row=total_row, column=1, value='TOTALS:').font = Font(bold=True)
            ws.cell(row=total_row, column=5, value=gross_qty).font = Font(bold=True)
            ws.cell(row=total_row, column=6, value=returned_qty).font = Font(bold=True)
            ws.cell(row=total_row, column=7, value=net_qty).font = Font(bold=True)
            ws.cell(row=total_row, column=8, value=f'৳{gross_amount:,.2f}').font = Font(bold=True)
            ws.cell(row=total_row, column=9, value=f'৳{returned_amount:,.2f}').font = Font(bold=True)
            ws.cell(row=total_row, column=10, value=f'৳{net_amount:,.2f}').font = Font(bold=True)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter if hasattr(column[0], 'column_letter') else None
                if column_letter:
                    for cell in column:
                        try:
                            if hasattr(cell, 'value') and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Create response
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=customer_wise_report_{customer.name}_{start_date}_to_{end_date}.xlsx'

            return response




class ReportExcelView(LoginRequiredMixin, View):
    """View for exporting general sales reports to Excel."""
    
    def post(self, request, *args, **kwargs):
        """Export general sales report to Excel."""
        # Get filter parameters
        customer_id = request.POST.get('customer')
        product_id = request.POST.get('product')
        sales_employee_id = request.POST.get('sales_employee')
        
        # Set default start and end dates
        today = timezone.now().date()
        first_day_of_month = today.replace(day=1)
        
        start_date = request.POST.get('start_date', first_day_of_month)
        end_date = request.POST.get('end_date', today)
        
        # Query SalesOrderItem directly for unified item-level report format
        items = SalesOrderItem.objects.select_related(
            'sales_order__customer',
            'sales_order__sales_employee',
            'product'
        )
        
        # Apply filters on item-level
        if customer_id:
            items = items.filter(sales_order__customer_id=customer_id)
        if product_id:
            items = items.filter(product_id=product_id)
        if sales_employee_id:
            items = items.filter(sales_order__sales_employee_id=sales_employee_id)
        if start_date and end_date:
            items = items.filter(sales_order__order_date__range=[start_date, end_date])
        
        # Order results by date (descending) and sales order id
        items = items.order_by('-sales_order__order_date', 'sales_order__id')
        
        # Query ReturnSalesItem data
        returns_qs = ReturnSalesItem.objects.select_related(
            'sales_order_item__sales_order__customer',
            'sales_order_item__sales_order__sales_employee',
            'sales_order_item__product',
            'return_sales'
        )
        
        # Apply same filters to returns
        if customer_id:
            returns_qs = returns_qs.filter(
                sales_order_item__sales_order__customer_id=customer_id
            )
        if product_id:
            returns_qs = returns_qs.filter(
                sales_order_item__product_id=product_id
            )
        if sales_employee_id:
            returns_qs = returns_qs.filter(
                sales_order_item__sales_order__sales_employee_id=sales_employee_id
            )
        if start_date and end_date:
            returns_qs = returns_qs.filter(
                return_sales__return_date__range=[start_date, end_date]
            )
        
        # Aggregate total returned quantity and amount
        returns_aggregated = returns_qs.aggregate(
            total_returned_qty=Sum('quantity'),
            total_returned_amount=Sum('total')
        )
        
        # Handle None values from aggregate (default to 0)
        returned_qty = returns_aggregated['total_returned_qty'] or 0
        returned_amount = returns_aggregated['total_returned_amount'] or 0
        
        # Group returns by sales_order_item_id
        returns_by_item = returns_qs.values('sales_order_item_id').annotate(
            returned_qty=Sum('quantity'),
            returned_amount=Sum('total')
        )
        
        # Build dictionary for quick lookup
        returns_dict = {
            r['sales_order_item_id']: {
                'qty': r['returned_qty'] or 0,
                'amount': r['returned_amount'] or 0
            }
            for r in returns_by_item
        }
        
        # Attach return data and calculate net values for each item
        items_with_data = []
        for item in items:
            # Attach return data for this specific item
            returns = returns_dict.get(item.id, {'qty': 0, 'amount': 0})
            item.returned_qty = returns['qty']
            item.returned_amount = returns['amount']
            
            # Calculate gross amount (before item-level discount)
            item.gross_amount = item.quantity * item.unit_price
            
            # Calculate net values for this item
            # Use item.total directly (already has item-level discount applied)
            item.net_qty = item.quantity - item.returned_qty
            item.net_amount = item.total - item.returned_amount
            
            items_with_data.append(item)
        
        # Calculate gross_qty from items queryset
        gross_qty = sum(item.quantity for item in items_with_data)
        
        # Calculate net_qty = gross_qty - returned_qty
        net_qty = gross_qty - returned_qty
        
        # Calculate gross_amount from items (using item.total which has item-level discount)
        gross_amount = sum(item.total for item in items_with_data)
        
        # Calculate net_amount = gross_amount - returned_amount
        net_amount = gross_amount - returned_amount
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "General Sales Report"
        
        # Add title
        ws.merge_cells('A1:J1')
        title_cell = ws.cell(row=1, column=1, value="General Sales Report")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add filter information
        filter_info = []
        if customer_id:
            try:
                customer = CustomerVendor.objects.get(id=customer_id)
                filter_info.append(f"Customer: {customer.name}")
            except CustomerVendor.DoesNotExist:
                pass
        if product_id:
            try:
                product = Product.objects.get(id=product_id)
                filter_info.append(f"Product: {product.name}")
            except Product.DoesNotExist:
                pass
        if sales_employee_id:
            try:
                employee = SalesEmployee.objects.get(id=sales_employee_id)
                filter_info.append(f"Sales Employee: {employee.full_name}")
            except SalesEmployee.DoesNotExist:
                pass
        if start_date and end_date:
            filter_info.append(f"Date Range: {start_date} to {end_date}")
        
        if filter_info:
            ws.merge_cells('A2:J2')
            filter_cell = ws.cell(row=2, column=1, value=" | ".join(filter_info))
            filter_cell.alignment = Alignment(horizontal="center")
            header_row = 4
        else:
            header_row = 3
        
        # Add headers with six columns
        headers = [
            'Date', 'Sales Order No', 'Customer Name', 'Product Name',
            'Gross Sold', 'Qty Returned', 'Net Qty',
            'Gross Amount', 'Return Amount', 'Net Amount'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        row_num = header_row + 1
        
        for item in items_with_data:
            row = [
                item.sales_order.order_date.strftime('%d/%m/%Y') if item.sales_order.order_date else '',
                item.sales_order.id,
                item.sales_order.customer.name if item.sales_order.customer else '',
                item.product.name if item.product else '',
                item.quantity,  # Gross Sold
                item.returned_qty or 0,  # Qty Returned
                item.net_qty,  # Net Qty
                float(item.total),  # Gross Amount (item.total has item-level discount)
                float(item.returned_amount or 0),  # Return Amount
                float(item.net_amount)  # Net Amount
            ]
            
            for col, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col, value=value)
            
            row_num += 1
        
        # Add totals row using aggregated values
        total_row = row_num + 1
        ws.merge_cells(f'A{total_row}:D{total_row}')
        ws.cell(row=total_row, column=1, value='TOTALS:').font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=gross_qty).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=returned_qty).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=net_qty).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=float(gross_amount)).font = Font(bold=True)
        ws.cell(row=total_row, column=9, value=float(returned_amount)).font = Font(bold=True)
        ws.cell(row=total_row, column=10, value=float(net_amount)).font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter if hasattr(column[0], 'column_letter') else None
            if column_letter:
                for cell in column:
                    try:
                        if hasattr(cell, 'value') and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=general_sales_report_{start_date}_to_{end_date}.xlsx'
        
        return response


class ReportPrintView(LoginRequiredMixin, View):
    """View for printing general sales reports."""
    template_name = 'reports/report_print.html'
    
    def post(self, request, *args, **kwargs):
        """Display printable general sales report."""
        # Get filter parameters
        customer_id = request.POST.get('customer')
        product_id = request.POST.get('product')
        sales_employee_id = request.POST.get('sales_employee')
        
        # Set default start and end dates
        today = timezone.now().date()
        first_day_of_month = today.replace(day=1)
        
        start_date = request.POST.get('start_date', first_day_of_month)
        end_date = request.POST.get('end_date', today)
        
        # Query SalesOrderItem directly for unified item-level report format
        items = SalesOrderItem.objects.select_related(
            'sales_order__customer',
            'sales_order__sales_employee',
            'product'
        )
        
        # Apply filters on item-level
        if customer_id:
            items = items.filter(sales_order__customer_id=customer_id)
        if product_id:
            items = items.filter(product_id=product_id)
        if sales_employee_id:
            items = items.filter(sales_order__sales_employee_id=sales_employee_id)
        if start_date and end_date:
            items = items.filter(sales_order__order_date__range=[start_date, end_date])
        
        # Order results by date (descending) and sales order id
        items = items.order_by('-sales_order__order_date', 'sales_order__id')
        
        # Query ReturnSalesItem data
        returns_qs = ReturnSalesItem.objects.select_related(
            'sales_order_item__sales_order__customer',
            'sales_order_item__sales_order__sales_employee',
            'sales_order_item__product',
            'return_sales'
        )
        
        # Apply same filters to returns
        if customer_id:
            returns_qs = returns_qs.filter(
                sales_order_item__sales_order__customer_id=customer_id
            )
        if product_id:
            returns_qs = returns_qs.filter(
                sales_order_item__product_id=product_id
            )
        if sales_employee_id:
            returns_qs = returns_qs.filter(
                sales_order_item__sales_order__sales_employee_id=sales_employee_id
            )
        if start_date and end_date:
            returns_qs = returns_qs.filter(
                return_sales__return_date__range=[start_date, end_date]
            )
        
        # Aggregate total returned quantity and amount
        returns_aggregated = returns_qs.aggregate(
            total_returned_qty=Sum('quantity'),
            total_returned_amount=Sum('total')
        )
        
        # Handle None values from aggregate (default to 0)
        returned_qty = returns_aggregated['total_returned_qty'] or 0
        returned_amount = returns_aggregated['total_returned_amount'] or 0
        
        # Group returns by sales_order_item_id
        returns_by_item = returns_qs.values('sales_order_item_id').annotate(
            returned_qty=Sum('quantity'),
            returned_amount=Sum('total')
        )
        
        # Build dictionary for quick lookup
        returns_dict = {
            r['sales_order_item_id']: {
                'qty': r['returned_qty'] or 0,
                'amount': r['returned_amount'] or 0
            }
            for r in returns_by_item
        }
        
        # Attach return data and calculate net values for each item
        items_with_data = []
        for item in items:
            # Attach return data for this specific item
            returns = returns_dict.get(item.id, {'qty': 0, 'amount': 0})
            item.returned_qty = returns['qty']
            item.returned_amount = returns['amount']
            
            # Calculate gross amount (before item-level discount)
            item.gross_amount = item.quantity * item.unit_price
            
            # Calculate net values for this item
            # Use item.total directly (already has item-level discount applied)
            item.net_qty = item.quantity - item.returned_qty
            item.net_amount = item.total - item.returned_amount
            
            items_with_data.append(item)
        
        # Calculate gross_qty from items queryset
        gross_qty = sum(item.quantity for item in items_with_data)
        
        # Calculate net_qty = gross_qty - returned_qty
        net_qty = gross_qty - returned_qty
        
        # Calculate gross_amount from items (using item.total which has item-level discount)
        gross_amount = sum(item.total for item in items_with_data)
        
        # Calculate net_amount = gross_amount - returned_amount
        net_amount = gross_amount - returned_amount
        
        # Get filter display names
        customer_name = None
        product_name = None
        employee_name = None
        
        if customer_id:
            try:
                customer = CustomerVendor.objects.get(id=customer_id)
                customer_name = customer.name
            except CustomerVendor.DoesNotExist:
                pass
        
        if product_id:
            try:
                product = Product.objects.get(id=product_id)
                product_name = product.name
            except Product.DoesNotExist:
                pass
        
        if sales_employee_id:
            try:
                employee = SalesEmployee.objects.get(id=sales_employee_id)
                employee_name = employee.full_name
            except SalesEmployee.DoesNotExist:
                pass
        
        # Format dates for display
        now = timezone.now()
        bdt = pytz.timezone('Asia/Dhaka')
        now_bdt = now.astimezone(bdt)
        
        context = {
            'report_items': items_with_data,
            'customer_name': customer_name,
            'product_name': product_name,
            'employee_name': employee_name,
            'start_date': start_date,
            'end_date': end_date,
            'gross_qty': gross_qty,
            'returned_qty': returned_qty,
            'net_qty': net_qty,
            'gross_amount': gross_amount,
            'returned_amount': returned_amount,
            'net_amount': net_amount,
            'user': request.user,
            'report_generated_at': now_bdt.strftime('%d/%m/%Y %I:%M %p'),
            'print_view': True
        }
        
        return render(request, self.template_name, context)
